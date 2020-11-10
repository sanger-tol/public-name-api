# coding: utf-8
# ToDo not implemented yet!
from __future__ import absolute_import

from swagger_server.test import BaseTestCase
from swagger_server.excel_utils import find_columns
from swagger_server.model import db, PnaSpecimen
from openpyxl import load_workbook

class TestCuratorsController(BaseTestCase):
    """CuratorsController integration test stubs"""

    def test_add_public_name(self):
        # No authorisation token given
        query_string = []
        response = self.client.open(
            '/public_name_api/public-name',
            method='PUT',
            query_string=query_string)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # Invalid authorisation token given
        query_string = []
        response = self.client.open(
            '/public_name_api/public-name',
            method='PUT',
            headers={"api-key": "12345678"},
            query_string=query_string)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # No taxonomyId given
        query_string = []
        response = self.client.open(
            '/public_name_api/public-name',
            method='PUT',
            headers={"api-key": self.api_key},
            query_string=query_string)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # No specimenId given
        query_string = [('taxonomyId', 6344)]
        response = self.client.open(
            '/public_name_api/public-name',
            method='PUT',
            headers={"api-key": self.api_key},
            query_string=query_string)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Taxonomy ID not in database
        query_string = [('taxonomyId', 999999999),
                         ('specimenId', 'SAN0000100')]
        response = self.client.open(
            '/public_name_api/public-name',
            method='PUT',
            headers={"api-key": self.api_key},
            query_string=query_string)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Taxonomy ID not correct for specimen
        query_string = [('taxonomyId', '9606'),
                        ('specimenId', 'SAN0000100')]
        response = self.client.open(
            '/public_name_api/public-name',
            method='PUT',
            headers={"api-key": self.api_key},
            query_string=query_string)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Specimen ID not in database - should create it
        query_string = [('taxonomyId', 6344),
                ('specimenId', 'SAN0000100xxxxx')]
        response = self.client.open(
            '/public_name_api/public-name',
            method='PUT',
            headers={"api-key": self.api_key},
            query_string=query_string)
        expect = [{
            "commonName": "lugworm",
            "family": "Arenicolidae",
            "genus": "Arenicola",
            "order": "None",
            "phylum": "Annelida",
            "prefix": "wuAreMari",
            "publicName": "wuAreMari2",
            "species": "Arenicola marina",
            "specimenId": "SAN0000100xxxxx",
            "taxaClass": "Polychaeta",
            "taxonomyId": 6344
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEquals(expect, response.json)

        # Specimen ID not in database and first for species - should create it
        query_string = [('taxonomyId', 9606),
                ('specimenId', 'SAN0000999xxxxx')]
        response = self.client.open(
            '/public_name_api/public-name',
            method='PUT',
            headers={"api-key": self.api_key},
            query_string=query_string)
        expect = [{
            "commonName": "human",
            "family": "Hominidae",
            "genus": "Homo",
            "order": "Primates",
            "phylum": "Chordata",
            "prefix": "mHomSap",
            "publicName": "mHomSap1",
            "species": "Homo sapiens",
            "specimenId": "SAN0000999xxxxx",
            "taxaClass": "Mammalia",
            "taxonomyId": 9606
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEquals(expect, response.json)

        # Existing - should return existing
        query_string = [('taxonomyId', 6344),
                ('specimenId', 'SAN0000100')]
        response = self.client.open(
            '/public_name_api/public-name',
            method='PUT',
            headers={"api-key": self.api_key},
            query_string=query_string)
        expect = [{
            "commonName": "lugworm",
            "family": "Arenicolidae",
            "genus": "Arenicola",
            "order": "None",
            "phylum": "Annelida",
            "prefix": "wuAreMari",
            "publicName": "wuAreMari1",
            "species": "Arenicola marina",
            "specimenId": "SAN0000100",
            "taxaClass": "Polychaeta",
            "taxonomyId": 6344
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEquals(expect, response.json)

    def test_add_species(self):
        # No authorisation token given
        query_string = []
        response = self.client.open(
            '/public_name_api/species',
            method='PUT',
            query_string=query_string)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # Invalid authorisation token given
        query_string = []
        response = self.client.open(
            '/public_name_api/species',
            method='PUT',
            headers={"api-key": "12345678"},
            query_string=query_string)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Taxonomy ID already in database
        body = {'taxonomyId': 6344,
                'species': 'Species',
                'prefix': 'whatever',
                'commonName': 'Common name', 
                'genus': 'Genus', 
                'family': 'Family', 
                'order': 'Order', 
                'taxaClass': 'Class', 
                'phylum': 'Phylum'}
        response = self.client.open(
            '/public_name_api/species',
            method='PUT',
            headers={"api-key": self.api_key},
            json=body)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Taxonomy ID not in database - should create it
        body = {'taxonomyId': 999999,
                'species': 'Species',
                'prefix': 'whatever',
                'commonName': 'Common name', 
                'genus': 'Genus', 
                'family': 'Family', 
                'order': 'Order', 
                'taxaClass': 'Class', 
                'phylum': 'Phylum'}
        response = self.client.open(
            '/public_name_api/species',
            method='PUT',
            headers={"api-key": self.api_key},
            json=body)
        expect = [{
            "commonName": "Common name",
            "family": "Family",
            "genus": "Genus",
            "order": "Order",
            "phylum": "Phylum",
            "prefix": "whatever",
            "species": "Species",
            "taxaClass": "Class",
            "taxonomyId": 999999
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEquals(expect, response.json)

        # Has it been added?
        query_string = [('taxonomyId', '999999')]
        response = self.client.open(
            '/public_name_api/species',
            method='GET',
            query_string=query_string)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEquals(expect, response.json)


    def test_edit_species(self):
        # No authorisation token given
        query_string = []
        response = self.client.open(
            '/public_name_api/species',
            method='PATCH',
            query_string=query_string)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # Invalid authorisation token given
        query_string = []
        response = self.client.open(
            '/public_name_api/species',
            method='PATCH',
            headers={"api-key": "12345678"},
            query_string=query_string)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Taxonomy ID not in database
        body = {'taxonomyId': 999999,
                'species': 'Species',
                'prefix': 'whatever',
                'commonName': 'Common name', 
                'genus': 'Genus', 
                'family': 'Family', 
                'order': 'Order', 
                'taxaClass': 'Class', 
                'phylum': 'Phylum'}
        response = self.client.open(
            '/public_name_api/species',
            method='PATCH',
            headers={"api-key": self.api_key},
            json=body)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Taxonomy ID in database - should edit it
        body = {'taxonomyId': 6344,
                'species': 'Species',
                'prefix': 'whatever',
                'commonName': 'Common name', 
                'genus': 'Genus', 
                'family': 'Family', 
                'order': 'Order', 
                'taxaClass': 'Class', 
                'phylum': 'Phylum'}
        response = self.client.open(
            '/public_name_api/species',
            method='PATCH',
            headers={"api-key": self.api_key},
            json=body)
        expect = [{
            "commonName": "Common name",
            "family": "Family",
            "genus": "Genus",
            "order": "Order",
            "phylum": "Phylum",
            "prefix": "whatever",
            "species": "Species",
            "taxaClass": "Class",
            "taxonomyId": 6344
        }]
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEquals(expect, response.json)

        # Has it changed?
        query_string = [('taxonomyId', '6344')]
        response = self.client.open(
            '/public_name_api/species',
            method='GET',
            query_string=query_string)
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEquals(expect, response.json)


    def test_validate_manifest(self):
        # No authorisation token given
        body = []
        response = self.client.open(
            '/public_name_api/validate-manifest',
            method='POST',
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/public_name_api/validate-manifest',
            method='POST',
            headers={"api-key": "12345678"},
            json=body)
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Excel file missing
        data = {}
        response = self.client.open(
            '/public_name_api/validate-manifest',
            method='POST',
            headers={"api-key": self.api_key},
            data=data)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Excel file with errors
        file = open('swagger_server/test/test-manifest-with-errors.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'), 
        }
        expected = {'errors': [{'message': 'Row 2: Expecting Arenicola marina, got Homo sapiens'},
            {'message': 'Row 3: Taxon ID 9999999 cannot be found'},
            {'message': 'Row 4: Genus only for Arenicola sp., not assigning public name'}]}
        response = self.client.open(
            '/public_name_api/validate-manifest',
            method='POST',
            headers={"api-key": self.api_key},
            data=data)
        file.close()
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEquals(expected, response.json)

        # Excel file correct
        file = open('swagger_server/test/test-manifest.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'), 
        }
        response = self.client.open(
            '/public_name_api/validate-manifest',
            method='POST',
            headers={"api-key": self.api_key},
            data=data)
        file.close()
        self.assert200(response, 'Not received a 200 response')
        self.assertEquals('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', response.content_type)

        # Save as Excel file
        file = open('swagger_server/test/test-manifest-validated.xlsx', 'wb')
        file.write(response.get_data())
        file.close()
        workbook = load_workbook(filename='swagger_server/test/test-manifest-validated.xlsx')
        sheet = workbook.active
        (taxon_id_column, specimen_id_column, scientific_name_column, public_name_column) = find_columns(sheet, "scientific_name")
        self.assertEquals('wuAreMari2', sheet.cell(row=2, column=public_name_column).value)
        self.assertEquals('wuAreMari3', sheet.cell(row=3, column=public_name_column).value)

        # Different column name for species
        file = open('swagger_server/test/test-manifest-col.xlsx', 'rb')
        data = {
            'excelFile': (file, 'test_file.xlsx'),
         }
        query_string = {'speciesColumnHeading': 'random column name'}
        response = self.client.open(
            '/public_name_api/validate-manifest',
            method='POST',
            headers={"api-key": self.api_key},
            query_string=query_string,
            data=data)
        file.close()
        self.assert200(response, 'Not received a 200 response')
        self.assertEquals('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', response.content_type)

        # Save as Excel file
        file = open('swagger_server/test/test-manifest-validated.xlsx', 'wb')
        file.write(response.get_data())
        file.close()
        workbook = load_workbook(filename='swagger_server/test/test-manifest-validated.xlsx')
        sheet = workbook.active
        (taxon_id_column, specimen_id_column, scientific_name_column, public_name_column) = find_columns(sheet, "random column name")
        self.assertEquals('wuAreMari2', sheet.cell(row=2, column=public_name_column).value)
        self.assertEquals('wuAreMari3', sheet.cell(row=3, column=public_name_column).value)


    def test_list_public_names(self):
        # Add a couple more specimens
        specimen2 = PnaSpecimen(specimen_id="SAN0000101", number=2, public_name="wuAreMari2")
        specimen2.species = self.species1
        specimen2.user = self.user1
        specimen3 = PnaSpecimen(specimen_id="SAN0000102", number=1, public_name="mHomSap1")
        specimen3.species = self.species2
        specimen3.user = self.user1
        db.session.add(specimen2)
        db.session.add(specimen3)
        db.session.commit()

        # No authorisation token given
        body = []
        response = self.client.open(
            '/public_name_api/public-name/all',
            method='GET',
            )
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/public_name_api/public-name/all',
            method='GET',
            headers={"api-key": "12345678"},
            )
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # No taxonomyId given
        query_string = []
        response = self.client.open(
            '/public_name_api/public-name/all',
            method='GET',
            headers={"api-key": self.api_key},
            query_string=query_string)
        expect = "wuAreMari1\tArenicola marina\tSAN0000100\t1\nwuAreMari2\tArenicola marina\tSAN0000101\t2\nmHomSap1\tHomo sapiens\tSAN0000102\t1"
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEquals('text/plain; charset=utf-8', response.content_type)
        self.assertEquals(expect, response.data.decode('utf-8'))

        # Taxonomy ID not in database
        query_string = [('taxonomyId', '999999999')]
        response = self.client.open(
            '/public_name_api/public-name/all',
            method='GET',
            headers={"api-key": self.api_key},
            query_string=query_string)
        self.assert400(response,
                       'Response body is : ' + response.data.decode('utf-8'))

        # Taxonomy ID given
        query_string = [('taxonomyId', '6344')]
        response = self.client.open(
            '/public_name_api/public-name/all',
            method='GET',
            headers={"api-key": self.api_key},
            query_string=query_string)
        expect = "wuAreMari1\tArenicola marina\tSAN0000100\t1\nwuAreMari2\tArenicola marina\tSAN0000101\t2"
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEquals('text/plain; charset=utf-8', response.content_type)
        self.assertEquals(expect, response.data.decode('utf-8'))

    def test_list_species(self):
        # No authorisation token given
        body = []
        response = self.client.open(
            '/public_name_api/species/all',
            method='GET',
            )
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # Invalid authorisation token given
        body = []
        response = self.client.open(
            '/public_name_api/species/all',
            method='GET',
            headers={"api-key": "12345678"},
            )
        self.assert401(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        # All correct
        query_string = []
        response = self.client.open(
            '/public_name_api/species/all',
            method='GET',
            headers={"api-key": self.api_key},
            query_string=query_string)
        expect = "wuAreMari\tArenicola marina\t6344\tlugworm\tArenicola\tArenicolidae\tNone\tPolychaeta\tAnnelida\nmHomSap\tHomo sapiens\t9606\thuman\tHomo\tHominidae\tPrimates\tMammalia\tChordata"
        self.assert200(response,
                       'Response body is : ' + response.data.decode('utf-8'))
        self.assertEquals('text/plain; charset=utf-8', response.content_type)
        self.assertEquals(expect, response.data.decode('utf-8'))

if __name__ == '__main__':
    import unittest
    unittest.main()
